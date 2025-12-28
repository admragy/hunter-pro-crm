"""
Report Generator Service - PDF & Excel Reports
Advanced reporting with charts and analytics
"""

import os
import io
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from PIL import Image
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend
import matplotlib.pyplot as plt
import seaborn as sns

logger = logging.getLogger(__name__)


class ReportService:
    """Advanced Report Generation"""
    
    def __init__(self):
        # Set style
        sns.set_style("whitegrid")
        plt.rcParams['font.family'] = 'DejaVu Sans'
        logger.info("✅ Report Service initialized")
    
    # ==================== PDF REPORTS ====================
    
    async def generate_pdf_report(
        self,
        title: str,
        data: Dict[str, Any],
        charts: Optional[List[Dict[str, Any]]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate comprehensive PDF report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            # Create PDF buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#6366f1'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Add title
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Add metadata
            if metadata:
                meta_text = f"<b>Generated:</b> {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}<br/>"
                meta_text += f"<b>Period:</b> {metadata.get('period', 'N/A')}<br/>"
                story.append(Paragraph(meta_text, styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
            
            # Add summary statistics
            if data.get('summary'):
                summary = data['summary']
                summary_data = [[k.replace('_', ' ').title(), str(v)] for k, v in summary.items()]
                
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0f172a')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.white)
                ]))
                
                story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
                story.append(Spacer(1, 0.1*inch))
                story.append(summary_table)
                story.append(Spacer(1, 0.3*inch))
            
            # Add charts
            if charts:
                for chart_config in charts:
                    chart_image = await self._generate_chart(chart_config)
                    if chart_image:
                        story.append(Paragraph(f"<b>{chart_config.get('title', 'Chart')}</b>", styles['Heading3']))
                        story.append(Spacer(1, 0.1*inch))
                        story.append(RLImage(chart_image, width=5*inch, height=3*inch))
                        story.append(Spacer(1, 0.3*inch))
            
            # Add detailed data table
            if data.get('details'):
                details = data['details']
                if details:
                    # Headers
                    headers = list(details[0].keys())
                    table_data = [headers]
                    
                    # Rows
                    for row in details:
                        table_data.append([str(row.get(h, '')) for h in headers])
                    
                    detail_table = Table(table_data)
                    detail_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1'))
                    ]))
                    
                    story.append(PageBreak())
                    story.append(Paragraph("<b>Detailed Data</b>", styles['Heading2']))
                    story.append(Spacer(1, 0.2*inch))
                    story.append(detail_table)
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(f"✅ PDF report generated: {len(pdf_bytes)} bytes")
            return pdf_bytes
            
        except Exception as e:
            logger.error(f"PDF generation error: {str(e)}")
            raise
    
    # ==================== EXCEL REPORTS ====================
    
    async def generate_excel_report(
        self,
        title: str,
        sheets: Dict[str, List[Dict[str, Any]]],
        metadata: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate Excel report with multiple sheets"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Process each sheet
            for sheet_name, data in sheets.items():
                if not data:
                    continue
                
                # Create DataFrame
                df = pd.DataFrame(data)
                
                # Create sheet
                ws = wb.create_sheet(title=sheet_name)
                
                # Add title
                ws['A1'] = title
                ws['A1'].font = Font(size=14, bold=True, color="6366f1")
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A1:' + chr(65 + len(df.columns) - 1) + '1')
                
                # Add metadata
                if metadata:
                    ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
                    ws['A3'] = f"Period: {metadata.get('period', 'N/A')}"
                
                # Add data
                start_row = 5 if metadata else 3
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Style header row
                        if r_idx == start_row:
                            cell.font = Font(bold=True, color="ffffff")
                            cell.fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(f"✅ Excel report generated: {len(excel_bytes)} bytes")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"Excel generation error: {str(e)}")
            raise
    
    # ==================== CHART GENERATION ====================
    
    async def _generate_chart(self, config: Dict[str, Any]) -> io.BytesIO:
        """Generate chart image"""
        try:
            chart_type = config.get('type', 'bar')
            data = config.get('data', {})
            title = config.get('title', '')
            
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_type == 'bar':
                ax.bar(data.keys(), data.values(), color='#6366f1')
            
            elif chart_type == 'line':
                ax.plot(list(data.keys()), list(data.values()), marker='o', color='#6366f1')
            
            elif chart_type == 'pie':
                ax.pie(data.values(), labels=data.keys(), autopct='%1.1f%%', startangle=90)
            
            ax.set_title(title, fontsize=14, fontweight='bold')
            ax.grid(True, alpha=0.3)
            plt.tight_layout()
            
            # Save to buffer
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
            buffer.seek(0)
            plt.close(fig)
            
            return buffer
            
        except Exception as e:
            logger.error(f"Chart generation error: {str(e)}")
            return None
    
    # ==================== DASHBOARD REPORT ====================
    
    async def generate_dashboard_report(
        self,
        period: str = "last_30_days"
    ) -> bytes:
        """Generate comprehensive dashboard PDF report"""
        # Mock data - replace with actual database queries
        data = {
            "summary": {
                "total_customers": 156,
                "active_deals": 42,
                "total_revenue": 84500,
                "win_rate": 65.3
            },
            "details": [
                {"customer": "John Doe", "deal_value": 5000, "status": "won"},
                {"customer": "Jane Smith", "deal_value": 3500, "status": "active"}
            ]
        }
        
        charts = [
            {
                "type": "bar",
                "title": "Monthly Revenue",
                "data": {"Jan": 25000, "Feb": 28000, "Mar": 31500}
            },
            {
                "type": "line",
                "title": "Customer Growth",
                "data": {"Jan": 120, "Feb": 138, "Mar": 156}
            }
        ]
        
        metadata = {
            "period": period,
            "generated_by": "Hunter Pro CRM"
        }
        
        return await self.generate_pdf_report(
            title="CRM Dashboard Report",
            data=data,
            charts=charts,
            metadata=metadata
        )


# Global service
report_service = ReportService()


async def get_report_service() -> ReportService:
    """Dependency injection"""
    return report_service
